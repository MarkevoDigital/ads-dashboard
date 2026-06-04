"""Gera a planilha de acessos (.xlsx) com os 20 clientes do dashboard."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = r"C:\Users\User\OneDrive\Documentos\Claude\Acessos_Dashboard_Markevo.xlsx"

ROWS = [
    ("NEP Objetivo", "nep", "bgIO7HpdVD", "1427648885223339", "212-147-9176"),
    ("Dr Carlos", "dr-carlos", "3yuWAD7KhX", "457463684925385", "915-240-4582"),
    ("Piping", "piping", "KLmz2ZLZHC", "834990480854318", "896-577-1923"),
    ("LimaX", "limax", "Kp9mWq3xLd", "779722861551815", ""),
    ("IPV7", "ipv7", "Rt6nZb8vHc", "313777577554670", ""),
    ("Bem me fiz", "bem-me-fiz", "Ya4cMs7kPe", "829671211248769", "537-488-7501"),
    ("Findrs", "findrs", "Wd2hQn9rTb", "", "692-574-8370"),
    ("Blank", "blank", "Lf8jXv3mGq", "1862390041356635", "153-430-5993"),
    ("JML", "jml", "Cs5pBk7wNd", "", "322-731-0364"),
    ("Zoemex", "zoemex", "Hm3tRy9zVx", "3380806325504088", ""),
    ("Celeiro", "celeiro", "Ne7qWd2cLp", "1091040991939408", "264-221-8233"),
    ("Proteplan", "proteplan", "Bx6vKm4hRs", "529893371378842", ""),
    ("Pantanal", "pantanal", "Tj9nGc3yWd", "", "443-676-8645"),
    ("Puríssima", "purissima", "Qa5rLp8mZk", "1397793118702309", ""),
    ("Reical", "reical", "Vh2bNs7xKd", "1480266830161455", "776-574-4113"),
    ("MT Leite", "mt-leite", "Dy8wCf4pMt", "1079892787307522", "965-181-2631"),
    ("Donaldson", "donaldson", "Gk3mRq9vLb", "", "240-677-0427"),
    ("Ápice", "apice", "Sn6cTx2hWd", "", "400-835-1194"),
    ("Servitec", "servitec", "Pf7jYm5kBq", "", "121-908-6011"),
    ("La Mullen", "la-mullen", "Zr4nLd8wKc", "1312573597377262", ""),
]
HEADERS = ["Cliente", "Login", "Senha", "Conta Meta", "Conta Google"]

wb = Workbook()
ws = wb.active
ws.title = "Acessos"

ARIAL = "Arial"
title_fill = PatternFill("solid", fgColor="1F2A44")
head_fill = PatternFill("solid", fgColor="2F4170")
alt_fill = PatternFill("solid", fgColor="EEF2FA")
thin = Side(style="thin", color="C9D3E6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Titulo
ws.merge_cells("A1:E1")
t = ws["A1"]
t.value = "Acessos — Dashboard Markevo  •  https://dashboard.markevo.com.br"
t.font = Font(name=ARIAL, size=13, bold=True, color="FFFFFF")
t.fill = title_fill
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 26

# Subtitulo / nota
ws.merge_cells("A2:E2")
n = ws["A2"]
n.value = "Login e senha diferenciam maiúsculas/minúsculas. Cada cliente vê apenas as próprias contas."
n.font = Font(name=ARIAL, size=9, italic=True, color="5A6577")
n.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 18

# Cabecalho
hrow = 3
for col, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=hrow, column=col, value=h)
    c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
    c.fill = head_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
ws.row_dimensions[hrow].height = 20

# Dados
for i, row in enumerate(ROWS):
    r = hrow + 1 + i
    for col, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=col, value=val if val else "—")
        c.font = Font(name=ARIAL, size=10,
                      color="000000",
                      bold=(col == 2))  # login em negrito
        c.alignment = Alignment(
            horizontal="left" if col == 1 else "center", vertical="center")
        c.border = border
        c.number_format = "@"  # texto (preserva IDs longos)
        if i % 2 == 1:
            c.fill = alt_fill

ws.freeze_panes = "A4"
widths = {"A": 18, "B": 14, "C": 14, "D": 20, "E": 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(OUT)
print("Planilha gerada:", OUT)
